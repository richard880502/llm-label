import io
import json
from pathlib import Path

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None

from ..database import get_db

router = APIRouter()


@router.get("/{project_id}/export")
def export_xlsx(project_id: int):
    if Workbook is None:
        raise HTTPException(500, "伺服器未安裝 openpyxl")

    with get_db() as conn:
        proj = conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
        if not proj:
            raise HTTPException(404, "Project not found")

        rows = conn.execute(
            "SELECT * FROM rows WHERE project_id=? ORDER BY source_row_number ASC",
            (project_id,),
        ).fetchall()

    if not rows:
        raise HTTPException(400, "No rows to export")

    first_original = json.loads(rows[0]["original_data"])
    original_cols = list(first_original.keys())
    review_cols = [
        "REVIEW_STATUS",
        "FINAL_RELEVANCE",
        "FINAL_LABELS",
        "FINAL_EMOTIONAL_SUBTYPES",
        "CORRECTED_RELEVANCE",
        "CORRECTED_LABELS",
        "CORRECTED_EMOTIONAL_SUBTYPES",
        "REVIEWER_NOTE",
        "REVIEWED_AT",
    ]
    all_cols = original_cols + review_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "複查結果"

    # Header style
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    review_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.fill = review_fill if col_name in review_cols else header_fill

    # Data rows
    for row in rows:
        original = json.loads(row["original_data"])
        final_relevance = row["corrected_relevance"] or row["ai_relevance"] or ""
        final_labels = _flatten_list(row["corrected_labels"] or row["ai_labels"])
        final_subtypes = _flatten_list(row["corrected_emotional_subtypes"] or row["ai_emotional_subtypes"])

        review_data = {
            "REVIEW_STATUS": row["status"],
            "FINAL_RELEVANCE": final_relevance,
            "FINAL_LABELS": final_labels,
            "FINAL_EMOTIONAL_SUBTYPES": final_subtypes,
            "CORRECTED_RELEVANCE": row["corrected_relevance"] or "",
            "CORRECTED_LABELS": _flatten_list(row["corrected_labels"]),
            "CORRECTED_EMOTIONAL_SUBTYPES": _flatten_list(row["corrected_emotional_subtypes"]),
            "REVIEWER_NOTE": row["reviewer_note"] or "",
            "REVIEWED_AT": row["reviewed_at"] or "",
        }
        full_row = {**original, **review_data}

        ws.append([full_row.get(c, "") for c in all_cols])

    # Column widths
    for col_idx, col_name in enumerate(all_cols, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        if col_name in ("CONTENT", "comment_content", "COMMENTS_CONTENT", "AI_RAW_RESPONSE",
                        "AI_LABEL_EVIDENCE_JSON", "AI_EMOTIONAL_EVIDENCE_JSON"):
            ws.column_dimensions[col_letter].width = 40
        elif col_name in review_cols:
            ws.column_dimensions[col_letter].width = 20
        else:
            ws.column_dimensions[col_letter].width = 16

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    stem = Path(proj["filename"]).stem
    download_name = f"{stem}_reviewed.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )


def _flatten_list(val: str | None) -> str:
    if not val:
        return ""
    try:
        items = json.loads(val)
        if isinstance(items, list):
            return ", ".join(str(i) for i in items)
    except Exception:
        pass
    return val
